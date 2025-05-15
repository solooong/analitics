import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
import os
import gspread
from google.oauth2.service_account import Credentials
import xml.etree.ElementTree as ET
from lxml import etree

def parser_xml():
    # --- 1. Парсим продажи ---
    parser = etree.XMLParser(recover=True)
    tree = etree.parse('b02_purchases_2025_04_29_return_value.xml', parser)
    root = tree.getroot()
    sales_rows = []
    # Список ключей для чтения из purchase
    keys = ['operDay', 'shop', 'cash', 'shift', 'number', 'amount', 'discountAmount', 'fiscalDocNum', 'order']
    # Сопоставление для переименования
    rename_map = {'amount': 'amount_itogo'}
    print('Обработка общих продаж')
    for purchase in root.findall('.//purchase'):
        purchase_data = {}
        for key in keys:
            value = purchase.attrib.get(key)
            new_key = rename_map.get(key, key)
            purchase_data[new_key] = value 
        for pos in purchase.findall('.//positions/position'):
            pos_data = dict(pos.attrib)
            row = purchase_data.copy()
            row.update(pos_data)
            sales_rows.append(row)
    final_df = pd.DataFrame(sales_rows)
    # --- 2. Парсим скидки ---
    print('Обработка продаж включающие скидки на товары')

    tree2 = ET.parse('b02_loy_2025_04_29_return_value.xml')
    root2 = tree2.getroot()
    discount_rows = []
    for purchase in root2.findall('.//purchase'):
        purchase_data = {key: purchase.attrib.get(key) for key in [
            'shop', 'cash', 'shift', 'number', 'saletime'
        ]}
        for disc in purchase.findall('.//discount'):
            disc_data = dict(disc.attrib)
            row = purchase_data.copy()
            row.update(disc_data)
            discount_rows.append(row)
    final_df_disc = pd.DataFrame(discount_rows)

 
    final_df_disc.rename(columns={'saletime': 'operDay', 'goodCode': 'goodsCode', 'positionId':'order'}, inplace=True)
    if 'amount' in final_df_disc.columns:
        final_df_disc.drop(columns=['amount'], inplace=True)
    # --- 3. Преобразуем типы ---
    print('Корректируем данные в ячейках')

    for df in [final_df, final_df_disc]:
        # Дата
        df['operDay'] = df['operDay'].str[:10]  # Обрезаем до первых 10 символов
        df['operDay'] = pd.to_datetime(df['operDay'], errors='coerce').dt.strftime('%d-%m-%Y')
        # Числовые поля
        for col in ['shop', 'cash', 'shift', 'number', 'goodsCode', 'quantity', 'amount', 'amount_itogo', 'positionId'  , 'order'
                    ,'discountAmount','count','cost','nds','ndsSum','discountValue','costWithDiscount']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').astype('float')
    print('Объеденяем данные в один DataFrame')

    # --- 4. Объединяем по ключам ---
    key_fields = ['operDay', 'shop', 'cash', 'shift', 'number', 'goodsCode', 'order']
    merged = final_df.merge(
        final_df_disc,
        how='left',
        on=key_fields)
        
    # Сохраняем данные
    merged.to_excel('temp_of_parser.xlsx', index=False)
    print('Итоговый файл: temp_of_parser.xlsx')
    
    return merged

# Чистим базу от лишних данных. Переименовываем столбцы для корректного отображения
def clean_df(clean):
    print('Чистим базу от лишних данных. ')
    key_collum_for_drop = ['sum_discount_item_2','sum_discount_item','fiscalDocNum',
                           'shift','number','order',
                           'departNumber','barCode','nds','ndsSum',
                           'dateCommit','insertType',
                             'AdvertActGUID', 'card-number', 'advertType', 'quantity', 'barCode',
                             'isDiscountPurchase', 'AdvertActDiscountType']
    
    print('Переименовываем столбцы для корректного отображения')
    dictonary={'operDay':'Дата','shop':'Магазин','cash':'Касса',
               'shift':'Смена','number':'Номер чека','amount_itogo':'Сумма чека',
               'discountAmount':'Сумма скидки чека',
               'positionId':'Позция в чеке','goodsCode':'ID Товара:','barCode':'Штрих код',
               'count':'Количество товара','cost':'Цена товара','nds':'НДС','ndsSum':'Сумма НДС',
               'discountValue':'Размер скидки на товар','costWithDiscount':
               'Цена товара со скидкой','amount':'Сумма на товар со скидкой', 
               'AdvertActExternalCode':'Номер журнала МА' ,
              ' itogo_discount_item':'Итого сумма скидки, руб',
                 'kolichesytvo_tovara':'Количество проданного товара, шт',
                'sum_of_item':'Итого продаж товара, руб',
                'vsego_chevok_item':'Итого чеков товара, шт',
                'sum_of_discount':'Итого продаж акции,руб',
                'vsego_chekov_discount':'Итого чеков акции, шт',
                'sum_of_sale_shop':'Итого продаж магазина,руб', 'vsego_chekov_shop': 'Всего чеков магазин'
}
    clean_df=pd.DataFrame(clean)
    
    key_collum_for_drop = [col for col in key_collum_for_drop if col in clean_df.columns]
    clean_df.drop(columns=key_collum_for_drop, inplace=True)
    # Rename columns
    clean_df = clean_df[~clean_df['AdvertActExternalCode'].isin(['0'])]# Удаляем строки товара без акции

    clean_df.rename(columns=dictonary, inplace=True)
    print('Меняем местами столобцы для большего удобства')
    clean_df.to_excel('temp_of_format_finality.xlsx', index=False)
    print('Итоговый файл: temp_of_format_finality.xlsx')   
    return clean_df
    
def analitics_colums(analitic_df):

    # Удаление строк с определённой акцией
    df=pd.DataFrame(analitic_df)
    df = df[df['AdvertActExternalCode'] != 'SR10_59320322']
    print("Убираем акции округления 50 копеек")
    # Ключевые поля для объединения
    key_fields = ['operDay', 'shop', 'cash', 'shift', 'number', 'order',  'goodsCode']

    # Аналитика по каждому товару в акции
    sales_of_item = (
        df.groupby(['operDay', 'shop', 'goodsCode'])
        .agg(
            sum_of_item=('amount', 'sum'),
            vsego_chevok_item=('fiscalDocNum', 'nunique'),
            sum_discount_item=('discountValue', 'sum'),
            kolichesytvo_tovara=('quantity', 'sum')
        )
    )

    # Аналитика по чекам с акцией (без разделения по товарам)
    sales_of_discount = (
        df.groupby(['operDay', 'shop', 'AdvertActExternalCode'], as_index=False)
        .agg(
            sum_of_discount=('amount', 'sum'),
            vsego_chekov_discount=('fiscalDocNum', 'nunique') #исправлено на лен. было не верное количество
        )
    )

    # Общая аналитика по магазинам и датам (по чекам)
    sales_of_shop = (
        df.drop_duplicates(subset='fiscalDocNum', keep='first')
        .groupby(['shop', 'operDay'], as_index=False)
        .agg(
            sum_of_sale_shop=('amount_itogo', 'sum'),
            vsego_chekov_shop=('fiscalDocNum', 'nunique')
        )
    )

    # Merge: объединяем всё обратно в исходный DataFrame
    df = df.merge(sales_of_item, how='left', on=['operDay', 'shop', 'goodsCode'])
    df = df.merge(sales_of_discount, how='left', on=['operDay', 'shop', 'AdvertActExternalCode'])
    df = df.merge(sales_of_shop, how='left', on=['shop', 'operDay'])

    # Заполняем NaN нулями
    df = df.fillna(0)
    print('Расчитываем доли и объеденяем в один файл')
    # Расчёты новых полей с защитой от деления на 0
    df['Доля скидки в цене товара'] = np.where(
        df['sum_of_item'] > 0,
        df['sum_discount_item'] / df['sum_of_item'],
        0
    )

    df['Доля продаж по всем товарам акции'] = np.where(
        df['sum_of_sale_shop'] > 0,
        df['sum_of_discount'] / df['sum_of_sale_shop'],
        0
    )

    df['Доля продаж товара в акции'] = np.where(
        df['sum_of_discount'] > 0,
        df['sum_of_item'] / df['sum_of_discount'],
        0
    )

    df['Средняя цена товара акции'] = np.where(
        df['kolichesytvo_tovara'] > 0,
        df['sum_of_item'] / df['kolichesytvo_tovara'],
        0
    )

    # Сохраняем результат
    # df.to_excel('temp_of_analitic.xlsx', index=False)
    # print('Итоговый файл: temp_of_analitic.xlsx')   
#     file_path = save_data_to_excel(df)
#     create_charts(df)
# # Добавляем графики как новый лист в Excel
#     from openpyxl import load_workbook
#     with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
#         wb = writer.book
#         if 'Графики' not in wb.sheetnames:
#             wb.create_sheet('Графики')

#         # TODO: если нужно — можно добавить графики программно
#         pass
    return df
# Указать обязательно информацию о количестве акций в данный период и количестве товара в акции

# === Функция построения графиков ===
def create_charts(df, output_path='charts.xlsx'):
    with pd.ExcelWriter(output_path, engine='openpyxl',  mode='w') as writer:
        # Создаем лист "Графики"
        worksheet = writer.book.create_sheet(title="Графики")
       
        print('Строим графики')
        img_streams = []
        def add_chart(fig):
            img_data = BytesIO()
            fig.savefig(img_data, format='png', bbox_inches='tight')
            img_data.seek(0)  # Важно сбросить позицию в начале потока
            img_streams.append(img_data)
            img = XLImage(img_data)
            worksheet.add_image(img)
                # Значение кода, которое нужно исключить (замените на ваше)
        exclude_code = "SR10_1485"  # Например, код билета докупателя
        # 1. Доля товара в акции (в цене)
        plt.figure(figsize=(8, 5))
        item_share = df.groupby('goodsCode')['amount'].sum() / df.groupby('goodsCode')['cost'].sum()
        item_share.dropna().sort_values(ascending=False).head(10).plot(kind='bar')
        plt.title("Доля товара в акции (в цене)")
        plt.ylabel("Доля (%)")
        add_chart(plt.gcf())
        plt.close()

        # 2. Доли акции в обороте (в цене)
        plt.figure(figsize=(8, 5))
        exclude_code = "SR10_1485"
        df_1 = df[df['AdvertActExternalCode'] != exclude_code]
        sales_by_act = df_1.groupby('AdvertActExternalCode')['amount'].sum()
        total_sales = sales_by_act.sum()
        (sales_by_act / total_sales).sort_values(ascending=False).head(15).plot(kind='pie', autopct='%1.1f%%')
        plt.title("Доли акции в обороте (в цене) - без билета Докупателя")
        add_chart(plt.gcf())
        plt.close()

        # 3. Популярность акции (по количеству чеков)
        exclude_code = "SR10_1485"
        # Фильтруем DataFrame, исключая указанный код
        filtered_df = df[df['AdvertActExternalCode'] != exclude_code]
        # Строим график на отфильтрованных данных
        plt.figure(figsize=(8, 5))
        popularity = filtered_df.groupby('AdvertActExternalCode')['fiscalDocNum'].nunique()
        popularity.sort_values(ascending=False).head(10).plot(kind='barh', color='lightgreen')
        plt.title("Популярность акции (по номеру) - Исключая билет докупателя")
        plt.xlabel("Количество уникальных чеков")
        add_chart(plt.gcf())
        plt.close()

        # 4. Сумма скидок по акциям4
        exclude_code = "SR10_1485"
        df_2 = df[df['AdvertActExternalCode'] != exclude_code]
        plt.figure(figsize=(8, 5))
        discount_sum = df_2.groupby('AdvertActExternalCode')['discountValue'].sum()
        discount_sum.sort_values(ascending=False).head(10).plot(kind='bar', color='salmon')
        plt.title("Сумма скидок по акциям (руб.)")
        plt.ylabel("Сумма скидок")
        add_chart(plt.gcf())
        plt.close()

        # Обновляем позиции графиков
        # writer._save()

# === Функция сохранения данных в Excel с разбивкой по магазинам ===
def save_data_to_excel(df, filename='output_data.xlsx'):
    print('Сохраняем данные в один файл.')
    print('Структуритуем по листам.')
    
    with pd.ExcelWriter(filename, engine='openpyxl', mode='w') as writer:
        # Лист 1: Все магазины
        df.to_excel(writer, sheet_name='Все магазины', index=False)

        # Листы по магазинам
        for shop in df['shop'].unique():
            df_shop = df[df['shop'] == shop]
            sheet_name = f"Магазин_{shop}"[:31]  # Ограничение длины названия листа
            df_shop.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"Данные сохранены в {filename}")
    return filename

# def math_of_colum():#  Тут расчитываем показатели вхождения 

def main():
    print("=== Шаг 1: Парсинг XML файлов ===")
    original_df = parser_xml()

    print("\n=== Шаг 2: Аналитика по продажам и акциям ===")
    analiz_df = analitics_colums(original_df)

    print("\n=== Шаг 3: Очистка и форматирование данных ===")
    df_cleaned_final = clean_df(analiz_df)
    
    print("\n=== Шаг 4: Построение графиков ===")
    create_charts(analiz_df)


 


    print("\n✅ Все этапы выполнены успешно!")


main()





# === Функция отправки файла в Google Sheets ===
def upload_to_google_sheets(file_path, sheet_name='Аналитика'):
    scope = ["https://spreadsheets.google.com/feeds ", "https://www.googleapis.com/auth/drive "]
    creds = Credentials.from_service_account_file('service_account.json', scopes=scope)
    client = gspread.authorize(creds)

    try:
        sh = client.open(sheet_name)
    except gspread.exceptions.SpreadsheetNotFound:
        sh = client.create(sheet_name)

    # Открываем первый лист и очищаем его
    worksheet = sh.sheet1
    worksheet.clear()

    # Загружаем данные из Excel
    df_all = pd.read_excel(file_path, sheet_name='Все магазины')
    data = [df_all.columns.values.tolist()] + df_all.values.tolist()
    worksheet.update(data)

    # Добавляем листы по магазинам
    xls = pd.ExcelFile(file_path)
    for sheet in xls.sheet_names:
        if sheet != "Все магазины" and sheet != "Графики":
            try:
                ws = sh.add_worksheet(title=sheet, rows="1000", cols="20")
            except gspread.exceptions.APIError:
                ws = sh.worksheet(sheet)
            df = pd.read_excel(file_path, sheet_name=sheet)
            data = [df.columns.values.tolist()] + df.values.tolist()
            ws.update(data)

    print(f"Файл загружен в Google Sheets: {sh.url}")

# === Основная функция аналитики ===