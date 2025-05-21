import pandas as pd
import numpy as np
import openpyxl
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
import os
import gspread
from google.oauth2.service_account import Credentials
import xml.etree.ElementTree as ET
from lxml import etree
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from copy import deepcopy

# from jinja2 import Environment, FileSystemLoader
# from weasyprint import HTML


current_date_str=datetime.now().strftime("%d_%m_%Y")
time = datetime.now().strftime("%d/%m_%H-%M")

# def generate_pdf_report(df, pivot_promo, pivot_products, charts_path='charts.xlsx', output_file='Отчет_по_акциям.pdf'):
#     template_str="""
#         <!DOCTYPE html>
#     <html>
#     <head>
#         <meta charset="utf-8">
#         <title>Аналитика</title>
#         <style>
#             body { font-family: sans-serif; padding: 20px; }
#             h1, h2 { text-align: center; color: #2c3e50; }
#             table { border-collapse: collapse; width: 100%; margin-bottom: 20px; }
#             th, td { border: 1px solid black; padding: 8px; text-align: left; }
#             img { max-width: 100%; height: auto; display: block; margin: 20px auto; }
#         </style>
#     </head>
#     <body>

#     <h1>Аналитика продаж и акций</h1>
#     <p style="text-align:center;">Дата: {{ date }}</p>

#     <h2>Свод по акциям</h2>
#     {{ promo_table }}

#     <h2>Свод по товарам</h2>
#     {{ product_table }}

#     <h2>Графики</h2>
#     {% for chart in charts %}
#     <img src="{{ chart }}" />
#     {% endfor %}

#     </body>
#     </html>
#     """
#     env = Environment(loader=FileSystemLoader('.'))
#     template = env.get_template('template.html')

#     # Преобразуем DataFrame в HTML
#     promo_html = pivot_promo.to_html(index=False, classes='table')
#     product_html = pivot_products.to_html(index=False, classes='table')

#     # Сохраняем временные изображения графиков
#     wb = load_workbook(charts_path)
#     chart_sheet = wb['Графики']
#     chart_paths = []

#     temp_dir = 'temp_charts'
#     os.makedirs(temp_dir, exist_ok=True)

#     for idx, img in enumerate(chart_sheet._images):
#         temp_img_path = os.path.join(temp_dir, f'chart_{idx}.png')
#         with open(temp_img_path, 'wb') as f:
#             f.write(img._data())
#         chart_paths.append(temp_img_path)

#     # Рендерим шаблон
#     html_out = template.render(
#         date=pd.Timestamp.now().strftime("%d-%m-%Y %H:%M"),
#         promo_table=promo_html,
#         product_table=product_html,
#         charts=chart_paths
#     )

#     # Сохраняем в PDF
#     HTML(string=html_out).write_pdf(output_file)

#     # Удаляем временные файлы
#     for path in chart_paths:
#         os.remove(path)
#     os.rmdir(temp_dir)

#     print(f"✅ Отчёт сохранён как {output_file}")

def parser_xml():
    # --- 1. Парсим продажи ---
    parser = etree.XMLParser(recover=True)
    tree = etree.parse('b02_purchases_2025_04_29_return_value.xml', parser)
    root = tree.getroot()
    sales_rows = []
    # Список ключей для чтения из purchase
    keys = ['operDay', 'shop', 'cash', 'shift', 'number', 'amount', 'discountAmount', 'fiscalDocNum', 'order']
    # Сопоставление для переименования
    rename_map = {'amount': 'amount_itogo'}
    print('Обработка общих продаж')
    for purchase in root.findall('.//purchase'):
        purchase_data = {}
        for key in keys:
            value = purchase.attrib.get(key)
            new_key = rename_map.get(key, key)
            purchase_data[new_key] = value 
        for pos in purchase.findall('.//positions/position'):
            pos_data = dict(pos.attrib)
            row = purchase_data.copy()
            row.update(pos_data)
            sales_rows.append(row)
    final_df = pd.DataFrame(sales_rows)
    # --- 2. Парсим скидки ---
    print('Обработка продаж включающие скидки на товары')
    tree2 = ET.parse('b02_loy_2025_04_29_return_value.xml')
    root2 = tree2.getroot()
    discount_rows = []
    for purchase in root2.findall('.//purchase'):
        purchase_data = {key: purchase.attrib.get(key) for key in [
            'shop', 'cash', 'shift', 'number', 'saletime'
        ]}
        for disc in purchase.findall('.//discount'):
            disc_data = dict(disc.attrib)
            row = purchase_data.copy()
            row.update(disc_data)
            discount_rows.append(row)
    final_df_disc = pd.DataFrame(discount_rows)
    final_df_disc.rename(columns={'saletime': 'operDay', 'goodCode': 'goodsCode', 'positionId':'order'}, inplace=True)
    if 'amount' in final_df_disc.columns:
        final_df_disc.drop(columns=['amount'], inplace=True)
    # --- 3. Преобразуем типы ---
    print('Корректируем данные в ячейках')

    for df in [final_df, final_df_disc]:
        # Дата
        df['operDay'] = df['operDay'].str[:10]  # Обрезаем до первых 10 символов
        df['operDay'] = pd.to_datetime(df['operDay'], errors='coerce').dt.strftime('%d-%m-%Y')
        # Числовые поля
        for col in ['shop', 'cash', 'shift', 'number', 'goodsCode', 'quantity', 'amount', 'amount_itogo', 'positionId'  , 'order'
                    ,'discountAmount','count','cost','nds','ndsSum','discountValue','costWithDiscount']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').astype('float')
    print('Объеденяем данные в один DataFrame')

    # --- 4. Объединяем по ключам ---
    key_fields = ['operDay', 'shop', 'cash', 'shift', 'number', 'goodsCode', 'order']
    merged = final_df.merge(
        final_df_disc,
        how='left',
        on=key_fields)
        
    # Меняем айди товара и айди акции на полноценные наименования
    df_name_of_discount = pd.read_excel('name_of_discount.xlsx')
    dictonary_rename=df_name_of_discount[['AdvertActExternalCode' , 'mechanik_of_discount', 'Название акции' ,'goodsCode' , 'Наименование товара']]
    keys_of_merge=['goodsCode','AdvertActExternalCode']
    merged=merged.merge(dictonary_rename,left_on=keys_of_merge, right_on=keys_of_merge, how='left' )
    
    merged.drop(columns='AdvertActExternalCode', inplace=True)
    merged.rename(columns={'Название акции' : 'AdvertActExternalCode' }, inplace=True)
    
    # Сохраняем данные
    # merged.to_excel(f'parser_{current_date_str}.xlsx', index=False)
    print(f'Объединение завершено {time}.xlsx')


    return merged

# Чистим базу от лишних данных. Переименовываем столбцы для корректного отображения
def clean_df(clean):
    print('Чистим базу от лишних данных. ')
    key_collum_for_drop = ['sum_discount_item_2','sum_discount_item',
                           'shift','number','order',
                           'departNumber','barCode','nds','ndsSum',
                           'dateCommit','insertType',
                             'AdvertActGUID', 'card-number', 'advertType', 'quantity', 'barCode',
                             'isDiscountPurchase', 'AdvertActDiscountType']
    
    print('Переименовываем столбцы для корректного отображения')
    dictonary={'operDay':'Дата','shop':'Магазин','cash':'Касса','mechanik_of_discount' :'Механика акции',
               'shift':'Смена','number':'Номер','amount_itogo':'Сумма чека',
               'discountAmount':'Сумма скидки чека',
               'positionId':'Позция в чеке','goodsCode':'ID Товара:','barCode':'Штрих код',
               'count':'Количество товара в чеке','cost':'Цена товара (ВЦ) в чеке','nds':'НДС','ndsSum':'Сумма НДС',
               'discountValue':'Размер скидки на товар в чеке','costWithDiscount':
               'Цена товара со скидкой в чеке','amount':'Итого товар со скидкой в чеке', 
               'AdvertActExternalCode':'Наименование акции' ,
              ' itogo_discount_item':'Итого сумма скидки, руб',
                 'kolichesytvo_tovara':'Количество проданного товара, шт',
                'sum_of_item':'Итого проданного товара, руб',
                'vsego_chevok_item':'Всего чеков товара, шт',
                'sum_of_discount':'Сумма продаж акции,руб','fiscalDocNum' : 'Номер чека',
                'vsego_chekov_discount':'Всего чеков акции, шт',
                'sum_of_sale_shop':'Сумма продаж магазина,руб', 'vsego_chekov_shop': 'Всего чеков магазин'
}
    clean_df=pd.DataFrame(clean)
    key_collum_for_drop = [col for col in key_collum_for_drop if col in clean_df.columns]
    clean_df.drop(columns=key_collum_for_drop, inplace=True)
    # Rename columns
    clean_df.rename(columns=dictonary, inplace=True)
    print('Меняем местами столобцы для большего удобства')
    clean_df=clean_df[['Дата','Магазин',
                       'Наименование акции','Механика акции','Номер чека', 'Сумма чека','ID Товара:',
                       'Наименование товара','Количество товара в чеке',
                       'Цена товара (ВЦ) в чеке','Размер скидки на товар в чеке',
                       'Цена товара со скидкой в чеке','Итого товар со скидкой в чеке',
                       'Итого проданного товара, руб','Всего чеков товара, шт',
                       'Количество проданного товара, шт','Сумма продаж акции,руб',
                       'Всего чеков акции, шт','Сумма продаж магазина,руб',
                        'Всего чеков магазин','Доля скидки в цене товара',
                        'Доля продаж всех товаров акции',
                        'Доля продаж товара в акции','Средняя цена товара акции']]
    
    
    clean_df.to_excel('_Промежуточный итог форматирования.xlsx', index=False)
    print('Итоговый файл: _Промежуточный итог форматирования.xlsx')   
    return clean_df
    
def analitics_colums(analitic_df):
    # Удаление строк с определённой акцией
    df=pd.DataFrame(analitic_df)
    df = df[df['AdvertActExternalCode'] != 'Округление суммы чека до 0,50 руб наличный расчёт']
    print("Убираем акции округления 50 копеек")
    # Ключевые поля для объединения
    key_fields = ['operDay', 'shop', 'cash', 'shift', 'number', 'order',  'goodsCode']
    # Аналитика по каждому товару в акции
    sales_of_item = (df.groupby(['operDay', 'shop', 'goodsCode'])
        .agg(
            sum_of_item=('amount', 'sum'),
            vsego_chevok_item=('fiscalDocNum', 'nunique'),
            sum_discount_item=('discountValue', 'sum'),
            kolichesytvo_tovara=('quantity', 'sum')))
    # Аналитика по чекам с акцией (без разделения по товарам)
    sales_of_discount = (df.groupby(['operDay', 'shop', 'AdvertActExternalCode'], as_index=False)
        .agg(
            sum_of_discount=('amount', 'sum'),
            vsego_chekov_discount=('fiscalDocNum', 'nunique')))
    # Общая аналитика по магазинам и датам (по чекам)
    sales_of_shop = (df.drop_duplicates(subset='fiscalDocNum', keep='first')
        .groupby(['shop', 'operDay'], as_index=False)
        .agg(
            sum_of_sale_shop=('amount_itogo', 'sum'),
            vsego_chekov_shop=('fiscalDocNum', 'nunique')))
    # Merge: объединяем всё обратно в исходный DataFrame
    df = df.merge(sales_of_item, how='left', on=['operDay', 'shop', 'goodsCode'])
    df = df.merge(sales_of_discount, how='left', on=['operDay', 'shop', 'AdvertActExternalCode'])
    df = df.merge(sales_of_shop, how='left', on=['shop', 'operDay'])
    # Заполняем NaN нулями
    df = df.fillna(0)
    print('Расчитываем доли и объеденяем в один файл')
    # Расчёты новых полей с защитой от деления на 0
    df['Доля скидки в цене товара'] = np.where(
        df['sum_of_item'] > 0,
        df['sum_discount_item'] / df['sum_of_item'] ,0 )
    df['Доля продаж всех товаров акции'] = np.where(
        df['sum_of_sale_shop'] > 0,
        df['sum_of_discount'] / df['sum_of_sale_shop'], 0 )
    df['Доля продаж товара в акции'] = np.where(
        df['sum_of_discount'] > 0,
        df['sum_of_item'] / df['sum_of_discount'],0    )
    df['Средняя цена товара акции'] = np.where(
        df['kolichesytvo_tovara'] > 0,
        df['sum_of_item'] / df['kolichesytvo_tovara'], 0)
    # Сохраняем результат
    # df['Наименование товара'] = df['Наименование товара'].replace(
    # '0',  # Что заменить
    # 'Товар без акции')
    
    df.to_excel(f'Аналитика_общая_{current_date_str}.xlsx', index=False)
    print(f'Итоговый файл: Аналитика_общая_{current_date_str}.xlsx')   
    file_path = save_data_to_excel(df)
    create_charts(df)
# # Добавляем графики как новый лист в Excel
    from openpyxl import load_workbook
    with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
        wb = writer.book
        if f'Графики_{current_date_str}' not in wb.sheetnames:
            wb.create_sheet(f'Графики_{current_date_str}')
        pass
    return df

# === Функция построения графиков ===
def create_charts(df, output_path=f'Графики_{current_date_str}.xlsx'):    
   
   with pd.ExcelWriter(output_path, engine='openpyxl',  mode='w') as writer:

        # Создаем лист "Графики"
        worksheet = writer.book.create_sheet(title="Графики")   
        print('Строим графики')
        img_streams = []
        def add_chart(fig):
            img_data = BytesIO()
            fig.savefig(img_data, format='png', bbox_inches='tight')
            img_data.seek(0)  # Важно сбросить позицию в начале потока
            img_streams.append(img_data)
            img = XLImage(img_data)
            worksheet.add_image(img)
        # Значение кода, которое нужно исключить (замените на ваше)
        exclude_code = "Скидка по Билету Докупателя"  # Например, код билета докупателя
        # 2. Доли акции в обороте (в цене)
        df['goodsCode'] = df['goodsCode'].replace(
    '0','Товар без акции')
        df['AdvertActExternalCode'] = df['AdvertActExternalCode'].replace(
    '0','Без акции') 
        plt.figure(figsize=(8, 5))
        exclude_code = "Скидка по Билету Докупателя"
        df = df[df['AdvertActExternalCode'] != exclude_code]
        sales_by_act = df.groupby('AdvertActExternalCode')['amount'].sum()
        total_sales = sales_by_act.sum()
        (sales_by_act / total_sales).sort_values(ascending=False).head(15).plot(kind='pie', autopct='%1.1f%%')
        plt.title("Доли акции в обороте (в цене) - без билета Докупателя")
        add_chart(plt.gcf())
        plt.close()
        # 3. Популярность акции (по количеству чеков)
        exclude_code = "Скидка по Билету Докупателя"
        # Фильтруем DataFrame, исключая указанный код
        filtered_df = df[df['AdvertActExternalCode'] != exclude_code]
        # Строим график на отфильтрованных данных
        plt.figure(figsize=(8, 5))
        popularity = filtered_df.groupby('AdvertActExternalCode')['fiscalDocNum'].nunique()
        popularity.sort_values(ascending=False).head(10).plot(kind='barh', color='lightgreen')
        plt.title("Популярность акции (по номеру) - Исключая билет докупателя")
        plt.xlabel("Количество уникальных чеков")
        add_chart(plt.gcf())
        plt.close()
        # 4. Сумма скидок по акциям4
        df_2 = df[df['AdvertActExternalCode'] != exclude_code]
        plt.figure(figsize=(8, 5))
        discount_sum = df_2.groupby('AdvertActExternalCode')['discountValue'].sum()
        discount_sum.sort_values(ascending=False).head(10).plot(kind='bar', color='salmon')
        plt.title("Сумма скидок по акциям (руб.)")
        plt.ylabel("Сумма скидок")
        add_chart(plt.gcf())
        plt.close()

        print(f'Сохраняем в отдельный файл Графики_{current_date_str}.xlsx ')
        # Обновляем позиции графиков
    # writer._save()

def save_final_report(df, pivot_promo, pivot_products, charts_path=f'Графики_{current_date_str}', output=f'Итоговый отчёт МА_{current_date_str}.xlsx'):
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Сохраняем данные
        df['Наименование товара'] = df['Наименование товара'].fillna('Товар без акциино')
        df['Наименование акции'] = df['Наименование акции'].fillna('Без акции')
        df['Наименование товара'] = df['Наименование товара'].replace(0,'Товар без акции')
        df['Наименование акции'] = df['Наименование акции'].replace(0,'Без акции')
        df['Механика акции'] = df['Механика акции'].replace(0,'Механика не указана')
        df.to_excel(writer, sheet_name='Все данные', index=False)
        pivot_promo.to_excel(writer, sheet_name='Свод по акциям', index=False)
        pivot_products.to_excel(writer, sheet_name='Свод по товарам', index=False)
        # Загружаем книгу с графиками
        charts_wb = openpyxl.load_workbook(charts_path)
        # Копируем лист с графиками
        if f'Графики' in charts_wb.sheetnames:
            ws_dest = writer.book.create_sheet(f'Графики')
            ws_src = charts_wb[f'Графики']
            # Копируем изображения
            for img in ws_src._images:
                ws_dest.add_image(deepcopy(img))
    print(f"✅ Отчёт сохранён в {output}")

# === Функция сохранения данных в Excel с разбивкой по магазинам ===
def save_data_to_excel(df, filename=f'Аналитика_по_магазинам_{current_date_str}.xlsx'):
    print('Сохраняем данные в один файл.')
    print('Структуритуем по листам.')
    with pd.ExcelWriter(filename, engine='openpyxl', mode='w') as writer:
        # Лист 1: Все магазины
        df.to_excel(writer, sheet_name='Все магазины', index=False)
        # Листы по магазинам
        for shop in df['shop'].unique():
            df_shop = df[df['shop'] == shop]
            sheet_name = f"Магазин_{shop}"[:31]  # Ограничение длины названия листа
            df_shop.to_excel(writer, sheet_name=sheet_name, index=False)
    print(f"Данные сохранены в {filename}")
    return filename

def pivot_by_promotions(df):
    # Группировка по акциям
    print('Подготавливаем сводные таблицы по акциям')
    pivot = df.groupby(['AdvertActExternalCode','mechanik_of_discount'] , as_index=False).agg({
    'amount': 'sum',
    'fiscalDocNum': 'nunique',
    'discountValue': 'sum'    })
    
    # Переименование столбцов для удобства
    pivot.rename(columns={
    'AdvertActExternalCode': 'Код акции',
    'amount': 'Сумма продаж по акции',
    'fiscalDocNum': 'Число чеков',
    'discountValue': 'Общая сумма скидок', 'mechanik_of_discount':'Механика акции'
    }, inplace=True)
    return pivot


def pivot_by_products(df):
    # Группировка по товарам
    print('Подготавливаем сводные таблицы по товарам')
    pivot = df.groupby('Наименование товара', as_index=False).agg({
    'amount': 'sum',
    'quantity': 'sum',
    'Средняя цена товара акции': 'mean',
    'Доля скидки в цене товара': 'mean'        })
    # Округление значений
    pivot['Средняя цена товара акции'] = pivot['Средняя цена товара акции'].round(2)
    pivot['Доля скидки в цене товара'] = (pivot['Доля скидки в цене товара']).round(2)
    # Переименование столбцов
    pivot.rename(columns={
    
    'amount': 'Сумма продаж по товару',
    'quantity': 'Количество проданного',
    'Средняя цена товара акции': 'Средняя цена',
    'Доля скидки в цене товара': 'Доля скидки (%)'
    }, inplace=True)
    return pivot

# Главная функция
def main():
    print("=== Шаг 1: Парсинг XML файлов ===")
    original_df = parser_xml()
    print("\n=== Шаг 2: Аналитика по продажам и акциям ===")
    analiz_df = analitics_colums(original_df)
    print("\n=== Шаг 3: Очистка и форматирование данных ===")
    df_cleaned_final = clean_df(analiz_df)
    # print("\n=== Шаг 4: Построение графиков ===")
    # create_charts(df_cleaned_final)
        # Создание сводных таблиц
    print("\n=== Шаг 4: Создание сводных таблиц ===")
    pivot_promo = pivot_by_promotions(analiz_df)
    pivot_products = pivot_by_products(analiz_df)
    
    # Сохранение всех листов в один файл
    print('Сохранение всех листов в один файл')
    with pd.ExcelWriter(f'Итоговый отчёт МА_{current_date_str}.xlsx', engine='openpyxl') as writer:
        analiz_df.to_excel(writer, sheet_name='Все данные', index=False)
        pivot_promo.to_excel(writer, sheet_name='Свод по акциям', index=False)
        pivot_products.to_excel(writer, sheet_name='Свод по товарам', index=False)
    save_final_report(
        df_cleaned_final,
        pivot_promo,
        pivot_products,
        charts_path=f'Графики_{current_date_str}.xlsx',  # замените на актуальную дату
        output=f'Итоговый отчёт МА_{current_date_str}.xlsx'
    )
    # generate_pdf_report(df_cleaned_final, pivot_promo, pivot_products)
    
main()

# dfgg=pd.read_excel('temp_of_format_finality.xlsx')
# pivot_by_promotions(dfgg)
